# 我想大部分人都知道，通常一个程序员会具有的美德。当然了，有三种：懒惰、暴躁、傲慢。
# 一个人写的烂软件将会给另一个人带来一份全职工作。
# 一鼓作气，考研是一种忍耐！！！
from random import choice,randint
from openpyxl import Workbook,load_workbook
def generateRandomInformation(filename):
    workbook=Workbook()
    worksheet=workbook.worksheets[0]
    worksheet.append(['姓名','课程','成绩'])
    first='赵钱李孙'
    middle='伟昀琛东'
    last='坤艳志'
    subjects=('语文','数学','英语')
    for i in range(200):
        name=choice(first)
        if randint(1,100)>50:
            name=name+choice(middle)
        name=name+choice(last)
        worksheet.append([name,choice(subjects),randint(0,100)])
    workbook.save(filename)
def getResult(oldfile,newfile):
    result=dict()
    workbook=load_workbook(oldfile)
    worksheet=workbook.worksheets[0]

    for row in worksheet.rows:
        if row[0].value=='姓名':
            continue
        name,subject,grade=map(lambda cell:cell.value,row)
        t=result.get(name,{})
        f=t.get(subject,0)
        if grade>f:
            t[subject]=grade
            result[name]=t
    workbook1=Workbook()
    worksheet1=workbook1.worksheets[0]
    worksheet1.append(['姓名','课程','成绩'])
    for name,t in result.items():
        print(name,t)
        for subject,grade in t.items():
            worksheet1.append([name,subject,grade])
    workbook1.save(newfile)
if __name__=='__main__':
    oldfile=r'd:\test.xlsx'
    newfile=r'd:\result.xlsx'
    generateRandomInformation(oldfile)
    getResult(oldfile,newfile)
