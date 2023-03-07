# 我想大部分人都知道，通常一个程序员会具有的美德。当然了，有三种：懒惰、暴躁、傲慢。
# 一个人写的烂软件将会给另一个人带来一份全职工作。
# 一鼓作气，考研是一种忍耐！！！

'''
1.什么叫模块_模块化编程的好处.安装扩展库openpyxl
2在/project/p_10目录下编写一个程序 p_10.py，输入参考
[root@master ~]# vi /opt/software/data/project/p_10/p_10.py
3.按i键进入编辑模式，输入以下代码
4、新建两个空文件
[root@master ~]# touch /opt/software/data/project/p_10/test.xlsx
[root@master ~]# touch /opt/software/data/project/p_10/result.xlsx;
5、然后运行程序
[root@master ~]# python3 /opt/software/data/project/p_10/p_10.py'''

#encoding=utf-8
from random import choice, randint
from openpyxl import Workbook, load_workbook
#生成随机数据
def generateRandomInformation(filename):
    workbook = Workbook()
    worksheet = workbook.worksheets[0]
    worksheet.append(['姓名', '课程', '成绩'])
    # 中文名字中的第一、第二、第三个字
    first = '周武郑王'
    middle = '明汉光美'
    last = '天杰致远'
    subjects = ('语文', '数学', '英语')
    for i in range(200):
        name = choice(first)
        # 按一定概率生成只有两个字的中文名字
        if randint(1, 100) > 50:
            name = name + choice(middle)
        name = name + choice(last)
        # 依次生成姓名、课程名称和成绩
        worksheet.append([name, choice(subjects), randint(0, 100)])
        # 保存数据，生成 Excel2007 格式的文件
        workbook.save(filename)
def getResult(oldfile, newfile):
    # 用于存放结果数据的字典
    result = dict()
    #打开原始数据
    workbook = load_workbook(oldfile)
    worksheet = workbook.worksheets[0]
    #遍历原始数据
    for row in worksheet.rows:
        if row[0].value == '姓名':
            continue
        # 姓名,课程名称,本次成绩
        name, subject, grade = map(lambda cell: cell.value, row)
        # 获取当前姓名对应的课程名称和成绩信息
# 如果 result 字典中不包含，则返回空字典
        t = result.get(name, {})
       # 获取当前学生当前课程的成绩，若不存在，返回 0
        f = t.get(subject, 0)
        # 只保留该学生该课程的最高成绩
        if grade > f:
            t[subject] = grade
            result[name] = t
    workbook1 = Workbook()
    worksheet1 = workbook1.worksheets[0]
    worksheet1.append(['姓名', '课程', '成绩'])

    # 将 result 字典中的结果数据写入 Excel 文件
    for name, t in result.items():
        print(name, t)
        for subject, grade in t.items():
            worksheet1.append([name, subject, grade])
    workbook1.save(newfile)

if __name__ == '__main__':
    oldfile = r'/opt/software/data/project/p_10/test.xlsx'     # 此处注意写正确你的测试文档路径
    newfile = r'/opt/software/data/project/p_10/result.xlsx'   # 此处注意写正确你的测试文档路径
    generateRandomInformation(oldfile)
    getResult(oldfile, newfile)
