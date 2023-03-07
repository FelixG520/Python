'''
Python上机1
实验一、Excel 文件成绩处理
实验目的：
1、了解扩展库 openpyxl 的安装与使用。
2、熟练运用字典结构解决实际问题。
实验内容：
假设某学校所有课程每学期允许多次考试，学生可随时参加考试，系统自动将每次成绩
添加到 Excel 文件（包含 3 列：姓名，课程，成绩）中，现期末要求统计所有学生每门课
程的最高成绩。
编写程序，模拟生成若干同学的成绩并写入 Excel 文件，其中学生姓名和课程名称均
可重复，也就是允许出现同一门课程的多次成绩，最后统计所有学生每门课程的最高成绩，
并写入新的 Excel 文件。

实验步骤：
1、在命令提示符环境使用 pip install openpyxl 命令安装扩展库 openpyxl。
2、编写代码。
'''

import openpyxl
import random


def autoGet(fileName):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(['姓名', '课程', '成绩'])

    first = tuple('一二三四')
    middle = tuple('五六七八')
    last = tuple('任榴圆')
    # 课程名称
    subjects = ('语文', '数学', '英语')
    for i in range(2000):
        line = []
        r = random.randint(1, 100)
        name = random.choice(first)
        # 按一定概率生成只有两个字的中文名字
        if r > 50:
            name = name + random.choice(middle)
        name = name + random.choice(last)
        # 依次生成姓名、课程名称和成绩
        line.append(name)
        line.append(random.choice(subjects))
        line.append(random.randint(0, 100))
        sheet.append(line)
        # 保存数据，生成excel
    wb.save(fileName)


def getMaxScore(oldfileName, newfilrName):
    wb = openpyxl.load_workbook(oldfileName)
    final_wb = openpyxl.Workbook()  # 创建一个新的excel工作簿对象用来存储学生各科目最高成绩

    start_sheet = wb.active  # 开始自动生成的文件
    final_sheet = final_wb.active
    final_sheet.title = '学生的各科成绩'
    final_sheet.append(['姓名', '课程', '最高分'])

    stuGrade = {}  # 创建一个记录全校学生成绩的字典

    for i in range(2, start_sheet.max_row + 1):
        stuName = start_sheet['A' + str(i)].value  # 获取学生名字
        lesName = start_sheet['B' + str(i)].value  # 获取课程名
        score = start_sheet['C' + str(i)].value  # 获取对应的分数

        # 判断stuGrade中是否存有该学生
        if stuName in stuGrade:
            stuGrade[stuName]
        else:
            stuGrade[stuName] = {}

        # 获取最大分数,先判断该学生下的成绩字典中是否存有这门课程
        if lesName in stuGrade[stuName]:
            if score > stuGrade[stuName][lesName]:
                stuGrade[stuName][lesName] = score
        else:
            stuGrade[stuName][lesName] = score

    for student in stuGrade:
        for lesson, score in stuGrade[student].items():
            final_sheet.append([student, lesson, score])

    final_wb.save(newfilrName)


if __name__ == '__main__':
    oldfile = '全校学生成绩.xlsx'
    newfile = '学生最终成绩表.xlsx'
    autoGet(oldfile)
    getMaxScore(oldfile, newfile)

