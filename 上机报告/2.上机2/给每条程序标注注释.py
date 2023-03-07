'''
Python程序设计上机2
实验目的：
1、熟悉程序的运行环境。
2、阅读指定的程序，并运行，然后给每条程序标注注释。
实验步骤：
1、Python IDLE环境里编辑程序，并运行。
2、给每条程序标注注释。
'''
from random import choice,randint
from openpyxl import Workbook,load_workbook
def generateRandomInformation(filename):
    workbook=Workbook()
    worksheet=workbook.worksheets[0]
    worksheet.append(['姓名','课程','成绩'])
    first='赵钱李孙'
    middle='伟昀琛东'
    last='坤艳志'
    subjects=('语文','数学','英语')    #课程名称
    for i in range(200):
        name=choice(first)
        if randint(1,100)>50:          #按一定概率生成只有两个字的中文名字
            name=name+choice(middle)
        name=name+choice(last)
        worksheet.append([name,choice(subjects),randint(0,100)])
    workbook.save(filename)            #保存数据
def getResult(oldfile,newfile):
    result=dict()                      #创建一个记录学生成绩的字典
    workbook=load_workbook(oldfile)
    worksheet=workbook.worksheets[0]

    for row in worksheet.rows:
        if row[0].value=='姓名':       #判断row中是否存有该学生
            continue
        name,subject,grade=map(lambda cell:cell.value,row)
        t=result.get(name,{})
        f=t.get(subject,0)
        if grade>f:                    #获取最大分数,先判断该学生下的成绩字典中是否存有这门课程
            t[subject]=grade
            result[name]=t
    workbook1=Workbook()
    worksheet1=workbook1.worksheets[0]
    worksheet1.append(['姓名','课程','成绩'])
    for name,t in result.items():
        print(name,t)
        for subject,grade in t.items():
            worksheet1.append([name,subject,grade])
    workbook1.save(newfile)
if __name__=='__main__':
    oldfile=r'd:\test.xlsx'
    newfile=r'd:\result.xlsx'
    generateRandomInformation(oldfile)
    getResult(oldfile,newfile)
